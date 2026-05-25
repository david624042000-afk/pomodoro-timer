"""
Export consolidated match results to a colour-coded Excel file.

Layout (one row per input term):
  No. | 輸入英文 | 匹配狀態 | 商品代碼 | 建議中文 | 資料庫英文 | 相似度 | 確認✓ | 備註

  • Single candidate  → one line per column
  • Multi-candidate   → ①②③ numbered, newline-separated in each column
  • Compound (split)  → grey header row, then indented ↳ sub-rows

Row colours:
  ✓ 完全對應  → light green    C6EFCE
  ~ 接近對應  → light yellow   FFEB9C
  ? 多項候選  → light orange   FFCC99
  ◆ LLM建議  → light blue     D9E1F2
  ○ 待手動   → light pink     FCE4D6
  💡 已拆分  → light grey     EDEDED
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .match import (
    STATUS_EXACT, STATUS_NEAR, STATUS_MULTI, STATUS_MANUAL, consolidate_results
)
from .llm import STATUS_LLM

STATUS_SPLIT = "💡 已拆分"

FILL = {
    STATUS_EXACT:  PatternFill("solid", fgColor="C6EFCE"),
    STATUS_NEAR:   PatternFill("solid", fgColor="FFEB9C"),
    STATUS_MULTI:  PatternFill("solid", fgColor="FFCC99"),
    STATUS_LLM:    PatternFill("solid", fgColor="D9E1F2"),
    STATUS_MANUAL: PatternFill("solid", fgColor="FCE4D6"),
    STATUS_SPLIT:  PatternFill("solid", fgColor="EDEDED"),
}
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SPLIT_FONT  = Font(bold=True, color="595959", italic=True)

COLUMNS = [
    ("No.",        6),
    ("輸入英文名稱",  42),
    ("匹配狀態",     12),
    ("商品代碼",     10),
    ("建議中文名稱",  24),
    ("資料庫英文",   42),
    ("相似度",       10),
    ("確認 ✓",       8),
    ("備註",        22),
]

NUMBERS = "①②③④⑤⑥⑦⑧"

_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LINE_H = 15   # Excel row-height units per line of text


def _row_height(values: list) -> float:
    max_lines = max((str(v).count("\n") + 1 for v in values if v), default=1)
    return max(18, max_lines * _LINE_H + 4)


def _write_header(ws):
    for ci, (hdr, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24


def _apply_row(ws, row_idx: int, values: list, fill: PatternFill,
               font=None, indent: bool = False):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill   = fill
        cell.border = _BORDER
        if font:
            cell.font = font
        cell.alignment = Alignment(
            horizontal="center" if ci == 1 else "left",
            vertical="center",
            wrap_text=True,
            indent=1 if (indent and ci == 2) else 0,
        )
    ws.row_dimensions[row_idx].height = _row_height(values)


def _candidates_to_cells(candidates: list[dict]) -> tuple:
    """Return (codes, chinese, db_english, notes) strings, newline-separated."""
    if not candidates:
        return "", "", "", ""
    if len(candidates) == 1:
        c = candidates[0]
        return c["code"], c["chinese"], c["db_english"], c["note"]
    codes, chinese, dben, notes = [], [], [], []
    for i, c in enumerate(candidates):
        prefix = NUMBERS[i] if i < len(NUMBERS) else f"({i+1})"
        codes.append(c["code"])
        chinese.append(f"{prefix}{c['chinese']}")
        dben.append(f"{prefix}{c['db_english']}")
        notes.append(c["note"])
    return "\n".join(codes), "\n".join(chinese), "\n".join(dben), "\n".join(notes)


def _build_export_rows(flat_rows: list[dict]) -> list[dict]:
    """
    Convert flat match results into export rows, grouping compound sub-terms
    under a parent header row.
    """
    consolidated = consolidate_results(flat_rows)

    # Separate compound sub-terms from standalone items
    parents: dict[str, list] = {}
    for item in consolidated:
        if item["is_sub"] and item["parent_en"]:
            parents.setdefault(item["parent_en"], []).append(item)

    # Rebuild order: walk consolidated list, emit compound header on first sub-term seen
    seen_parents: set = set()
    ordered: list = []   # ("single", item) | ("compound", parent_text, [sub_items])
    for item in consolidated:
        if item["is_sub"]:
            parent = item["parent_en"]
            if parent not in seen_parents:
                seen_parents.add(parent)
                ordered.append(("compound", parent, parents[parent]))
        else:
            ordered.append(("single", None, item))

    export_rows = []
    seq = 0
    for kind, parent, payload in ordered:
        seq += 1
        if kind == "single":
            codes, chinese, dben, notes = _candidates_to_cells(payload["candidates"])
            export_rows.append({
                "seq":             seq,
                "label":           payload["input_en"],
                "status":          payload["status"],
                "codes":           codes,
                "chinese":         chinese,
                "db_english":      dben,
                "notes":           notes,
                "is_split_header": False,
                "is_sub":          False,
            })
        else:
            sub_items = payload
            export_rows.append({
                "seq":             seq,
                "label":           parent,
                "status":          f"{STATUS_SPLIT} {len(sub_items)}項",
                "codes":           "",
                "chinese":         "",
                "db_english":      "",
                "notes":           "",
                "is_split_header": True,
                "is_sub":          False,
            })
            for sub in sub_items:
                codes, chinese, dben, notes = _candidates_to_cells(sub["candidates"])
                export_rows.append({
                    "seq":             seq,
                    "label":           f"↳ {sub['input_en']}",
                    "status":          sub["status"],
                    "codes":           codes,
                    "chinese":         chinese,
                    "db_english":      dben,
                    "notes":           notes,
                    "is_split_header": False,
                    "is_sub":          True,
                })
    return export_rows


def export_excel(flat_rows: list[dict], output_path=None) -> Path:
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path.cwd() / f"trademark_matches_{ts}.xlsx"
    output_path = Path(output_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商標比對結果"
    ws.freeze_panes = "A2"

    _write_header(ws)

    for ri, er in enumerate(_build_export_rows(flat_rows), start=2):
        # Determine fill — split header uses STATUS_SPLIT, others use exact status
        status = er["status"]
        if er["is_split_header"]:
            fill = FILL[STATUS_SPLIT]
        else:
            fill = next(
                (FILL[k] for k in FILL if status == k),
                FILL[STATUS_MANUAL]
            )
        font = SPLIT_FONT if er["is_split_header"] else None

        _apply_row(ws, ri, [
            er["seq"], er["label"], status,
            er["codes"], er["chinese"], er["db_english"], er["notes"],
            "", "",   # 確認 / 備註
        ], fill, font=font, indent=er["is_sub"])

    # ── Legend ────────────────────────────────────────────────────────────
    leg = wb.create_sheet("說明")
    legend_rows = [
        ("顏色", "狀態", "說明"),
        ("綠色",  STATUS_EXACT,        "完全對應，可直接確認"),
        ("黃色",  STATUS_NEAR,         "接近對應，請確認是否合適"),
        ("橘色",  STATUS_MULTI,        "多項候選（①②③），刪除不需要的列後保留正確項"),
        ("淡藍",  STATUS_LLM,          "AI 建議，請確認語意是否正確"),
        ("灰色",  STATUS_SPLIT+" N項", "複合詞條已自動拆分，子項縮排顯示"),
        ("粉紅",  STATUS_MANUAL,       "無法自動比對，請手動填寫"),
    ]
    fills_leg = [
        HEADER_FILL,
        FILL[STATUS_EXACT], FILL[STATUS_NEAR], FILL[STATUS_MULTI],
        FILL[STATUS_LLM], FILL[STATUS_SPLIT], FILL[STATUS_MANUAL],
    ]
    for r, (row_data, fl) in enumerate(zip(legend_rows, fills_leg), start=1):
        for c, val in enumerate(row_data, start=1):
            cell = leg.cell(row=r, column=c, value=val)
            cell.fill   = fl
            cell.border = _BORDER
            if r == 1:
                cell.font = HEADER_FONT
            leg.column_dimensions[get_column_letter(c)].width = 24

    wb.save(output_path)
    return output_path
